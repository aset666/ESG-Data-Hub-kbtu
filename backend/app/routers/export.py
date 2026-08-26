import io
from datetime import datetime

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session, joinedload

from app.auth import get_current_user
from app.database import get_db
from app.models import Metric, User

router = APIRouter(prefix="/api/export", tags=["export"])

HEADERS = [
    "Блок ESG", "Категория", "Метрика", "Определение", "Единица измерения",
    "Стандарты", "Scope", "Статус",
    "Тип источника", "Система", "Частота обновления", "Формат",
    "Подразделение", "Владелец данных", "Ответственный за сбор",
    "Email", "Телефон", "Мессенджер", "Уровень доступа",
    "Место хранения", "Дата последнего обновления",
    "Полнота (1-5)", "Точность (1-5)", "Своевременность (1-5)", "Проблемы",
]

STATUS_LABELS = {
    "collected": "Собирается",
    "partial": "Частично",
    "not_collected": "Не собирается",
    "planned": "Планируется",
}


@router.get("/excel")
def export_excel(db: Session = Depends(get_db), _user: User = Depends(get_current_user)):
    metrics = db.query(Metric).options(
        joinedload(Metric.source), joinedload(Metric.responsible), joinedload(Metric.storage_quality)
    ).order_by(Metric.block, Metric.category, Metric.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "ESG Data Landscape"

    ws.append(HEADERS)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for m in metrics:
        src = m.source
        resp = m.responsible
        sq = m.storage_quality
        ws.append([
            m.block.value, m.category, m.name, m.definition or "", m.unit or "",
            ", ".join(m.standards or []), m.scope or "", STATUS_LABELS.get(m.status.value, m.status.value),
            src.source_type.value if src else "", src.system_name if src else "",
            src.update_frequency.value if src else "", src.data_format.value if src else "",
            resp.department if resp else "", resp.data_owner if resp else "", resp.data_steward if resp else "",
            resp.contact_email if resp else "", resp.contact_phone if resp else "", resp.contact_messenger if resp else "",
            resp.access_level.value if resp else "",
            sq.location.value if sq else "", sq.last_updated.strftime("%Y-%m-%d") if sq and sq.last_updated else "",
            sq.quality_completeness if sq else "", sq.quality_accuracy if sq else "", sq.quality_timeliness if sq else "",
            ", ".join(sq.issues or []) if sq else "",
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"ESG_Data_Landscape_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
